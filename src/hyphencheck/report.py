"""Write the audit out as the spreadsheet the proofreader actually works from."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .audit import AuditResult
from .model import CONSISTENCY_RULE, RULE_TITLES, Break, Verdict

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)

FILLS = {
    Verdict.VIOLATION: PatternFill("solid", fgColor="F8CBAD"),
    Verdict.SUSPECT: PatternFill("solid", fgColor="FFE699"),
    Verdict.NEEDS_CHECK: PatternFill("solid", fgColor="FFF2CC"),
    Verdict.ADVISORY: PatternFill("solid", fgColor="DEEBF7"),
}

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RULE_COLUMNS = list(range(1, 10)) + [CONSISTENCY_RULE]


def _page(brk: Break) -> str:
    return brk.book_page or f"(pdf {brk.pdf_page})"


def _write_header(sheet, headers: list[str], row: int = 1) -> None:
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[row].height = 30


def _autosize(sheet, widths: dict[int, int]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width


def _context(brk: Break) -> str:
    return f"…{brk.line_text[-45:]} / {brk.next_line_text[:45]}…"


def _instance_row(brk: Break) -> list:
    row = [
        _page(brk),
        brk.pdf_page,
        brk.display,
        brk.word,
        brk.kind,
        brk.worst.value,
        brk.flagged_rules,
        brk.reason,
    ]
    for rule in RULE_COLUMNS:
        finding = brk.finding_for(rule)
        row.append(finding.verdict.value if finding else Verdict.NOT_APPLICABLE.value)
    row.append("; ".join(brk.notes))
    row.append(_context(brk))
    return row


INSTANCE_HEADERS = [
    "Page", "PDF page", "Break as set", "Word", "Type", "Verdict",
    "Rules flagged", "Reason",
] + [f"R{r}" if isinstance(r, int) else "Consist." for r in RULE_COLUMNS] + ["Notes", "Context"]

INSTANCE_WIDTHS = {1: 8, 2: 9, 3: 22, 4: 20, 5: 11, 6: 13, 7: 16, 8: 70}
INSTANCE_WIDTHS.update({i: 7 for i in range(9, 9 + len(RULE_COLUMNS))})
INSTANCE_WIDTHS[9 + len(RULE_COLUMNS)] = 34
INSTANCE_WIDTHS[10 + len(RULE_COLUMNS)] = 60


def _sort_key(brk: Break):
    page = brk.book_page
    return (int(page) if page.isdigit() else 10**6 + brk.pdf_page, brk.line_index)


def _add_instance_sheet(workbook, title: str, breaks: list[Break], note: str = "") -> None:
    sheet = workbook.create_sheet(title)
    start = 1
    if note:
        sheet.cell(row=1, column=1, value=note).font = SECTION_FONT
        start = 3
    _write_header(sheet, INSTANCE_HEADERS, row=start)
    for offset, brk in enumerate(sorted(breaks, key=_sort_key), start=start + 1):
        for column, value in enumerate(_instance_row(brk), start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(column in (8, len(INSTANCE_HEADERS))))
        fill = FILLS.get(brk.worst)
        if fill:
            sheet.cell(row=offset, column=6).fill = fill
    _autosize(sheet, INSTANCE_WIDTHS)
    sheet.freeze_panes = sheet.cell(row=start + 1, column=4)
    if len(breaks):
        sheet.auto_filter.ref = (
            f"A{start}:{get_column_letter(len(INSTANCE_HEADERS))}{start + len(breaks)}"
        )


def _add_summary(workbook, result: AuditResult) -> None:
    sheet = workbook.create_sheet("Summary", 0)
    sheet.cell(row=1, column=1, value="End-of-line hyphenation audit").font = TITLE_FONT
    rows = [
        ("File", Path(result.pdf_path).name),
        ("Audited", datetime.now().strftime("%d %B %Y, %H:%M")),
        ("Rule 1 source", result.dictionary_source),
        ("Merriam-Webster lookups this run", result.api_calls),
        ("Time taken", f"{result.elapsed:.0f} seconds"),
    ]
    row = 3
    for label, value in rows:
        sheet.cell(row=row, column=1, value=label).font = SECTION_FONT
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Totals").font = TITLE_FONT
    row += 1
    for label, value in result.counts().items():
        sheet.cell(row=row, column=1, value=label).font = SECTION_FONT
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Every break was checked against every rule").font = TITLE_FONT
    row += 1
    sheet.cell(
        row=row, column=1,
        value="Counts below are of breaks flagged by each rule. A break clearing one rule "
              "is still checked against all the others.",
    )
    row += 2
    _write_header(sheet, ["Rule", "Title", "Violations", "Needs check", "Advisories"], row=row)
    header_row = row
    row += 1
    for rule in RULE_COLUMNS:
        title = RULE_TITLES.get(rule, "Same word divided two different ways")
        counts = {Verdict.VIOLATION: 0, Verdict.NEEDS_CHECK: 0, Verdict.ADVISORY: 0}
        for brk in result.breaks:
            finding = brk.finding_for(rule)
            if not finding:
                continue
            verdict = finding.verdict
            if verdict is Verdict.SUSPECT:
                verdict = Verdict.NEEDS_CHECK
            if verdict in counts:
                counts[verdict] += 1
        label = f"Rule {rule}" if isinstance(rule, int) else rule
        for column, value in enumerate(
            [label, title, counts[Verdict.VIOLATION], counts[Verdict.NEEDS_CHECK],
             counts[Verdict.ADVISORY]],
            start=1,
        ):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = BORDER
        row += 1
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    row += 2
    sheet.cell(row=row, column=1, value="Every flagged item, by page").font = TITLE_FONT
    row += 2
    _write_header(sheet, ["Page", "Break as set", "Verdict", "Rules", "Reason"], row=row)
    row += 1
    for brk in sorted(result.flagged + result.advisories, key=_sort_key):
        values = [_page(brk), brk.display, brk.worst.value, brk.flagged_rules, brk.reason]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=column, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(column == 5))
        fill = FILLS.get(brk.worst)
        if fill:
            sheet.cell(row=row, column=3).fill = fill
        row += 1

    if not result.flagged and not result.advisories:
        sheet.cell(row=row, column=1, value="No violations found.")

    _autosize(sheet, {1: 12, 2: 24, 3: 14, 4: 20, 5: 95})


def _add_line_breaks(workbook, result: AuditResult) -> None:
    sheet = workbook.create_sheet("Line Breaks (Rule 6)")
    sheet.cell(
        row=1, column=1,
        value="Breaks between words — initials, regnal numerals, Jr./Sr. "
              "These are not hyphen breaks, so they are listed separately.",
    ).font = SECTION_FONT
    _write_header(sheet, ["Page", "PDF page", "Line ends", "Next line starts", "Verdict", "Reason"], row=3)
    for offset, finding in enumerate(result.line_breaks, start=4):
        values = [
            finding.book_page or f"(pdf {finding.pdf_page})",
            finding.pdf_page, finding.tail, finding.head,
            finding.verdict.value, finding.message,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(column == 6))
        fill = FILLS.get(finding.verdict)
        if fill:
            sheet.cell(row=offset, column=5).fill = fill
    if not result.line_breaks:
        sheet.cell(row=4, column=1, value="Nothing found.")
    _autosize(sheet, {1: 10, 2: 10, 3: 26, 4: 26, 5: 14, 6: 85})
    sheet.freeze_panes = sheet.cell(row=4, column=1)


def _add_rules_sheet(workbook) -> None:
    sheet = workbook.create_sheet("Rule Key")
    _write_header(sheet, ["Column", "Rule", "What it checks"], row=1)
    descriptions = {
        1: "The break lands on a division point in Merriam-Webster's dotted entry.",
        2: "At least two letters before the hyphen and three after it.",
        3: "A possessive 's is discounted when counting the three letters after.",
        4: "The word is not set tight against an em dash.",
        5: "A hyphenated compound is broken only at one of its own hyphens.",
        6: "Proper nouns: MW entry, then a recognizable morpheme, then a vowel.",
        7: "Where MW allows more than one point, the morpheme boundary is preferred.",
        8: "URLs, domains and email addresses never take a hyphen at a line break.",
        9: "Foreign-language words with no MW entry are flagged.",
        CONSISTENCY_RULE: "The same word is divided the same way throughout the book.",
    }
    for offset, rule in enumerate(RULE_COLUMNS, start=2):
        label = f"R{rule}" if isinstance(rule, int) else "Consist."
        name = RULE_TITLES.get(rule, "Internal consistency")
        for column, value in enumerate([label, name, descriptions[rule]], start=1):
            cell = sheet.cell(row=offset, column=column, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(column == 3))
    _autosize(sheet, {1: 10, 2: 46, 3: 80})


def write(result: AuditResult, output_path: str | Path) -> Path:
    """Write the four-tab workbook and return the path it was saved to."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    _add_summary(workbook, result)
    _add_instance_sheet(
        workbook, "Flagged Items", result.flagged + result.advisories,
        note="Every break that needs your attention, sorted by page.",
    )
    _add_instance_sheet(
        workbook, "All Instances", result.breaks,
        note="Every end-of-line hyphen in the proof, with each rule's verdict in "
             "columns R1–R9 (see the Rule Key tab).",
    )
    _add_line_breaks(workbook, result)
    _add_rules_sheet(workbook)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
