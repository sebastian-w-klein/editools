"""End to end: a proof PDF in, a spreadsheet out."""

from __future__ import annotations

import json

from openpyxl import load_workbook

from hyphencheck import audit, report
from hyphencheck.dictionary import Dictionary
from hyphencheck.model import Verdict
from conftest import build_cache


def run_audit(proof_pdf, tmp_path):
    cache = tmp_path / "mw-cache.json"
    cache.write_text(json.dumps(build_cache()), encoding="utf-8")
    dictionary = Dictionary(api_key="", cache_path=cache, offline=True)
    return audit.run(str(proof_pdf), dictionary)


def test_the_expected_violations_are_found(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    violations = {brk.display: brk.flagged_rules for brk in result.violations}
    expected = {
        "butch-/er's": "Rule 3",                    # possessive, cleared Rule 1
        "cher-/ries": "Rule 4",                     # em dash, cleared Rule 1
        "fold-/up": "Rule 2",                       # 2 letters after
        "photo-/grapher": "Rule 1",                 # MW: pho·tog·ra·pher
        "Ply-/mouth": "Rule 1",                     # MW: Plym·outh
        "Eng-/lish": "Rule 1",                      # MW: En·glish
        "co-/meth": "Rule 1",                       # MW carries no division point
        "pov-/erty-stricken": "Rule 5",             # broken inside a component
        "power-/ful": "Rule 4",                     # em dash before the word
        "www.exam-/ple.com": "Rule 8",              # hyphen added to a URL
        "Wordswor-/th": "Rule 2",                   # splits "-worth", 2 letters after
    }
    for display, rule in expected.items():
        assert display in violations, f"{display} was not flagged"
        assert rule in violations[display], f"{display}: expected {rule}, got {violations[display]}"


def test_valid_breaks_are_left_alone(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    clean = {brk.display for brk in result.clean}
    assert "cross-/legged" in clean        # a compound's own hyphen
    assert "poverty-/stricken" in clean    # the ruleset's own correct example


def test_an_invented_name_is_flagged_not_guessed(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    flagged = {brk.display: brk for brk in result.needs_check}
    assert "Mar-/volene" in flagged
    assert "Consistency" in flagged["Mar-/volene"].flagged_rules


def test_line_break_findings_cover_initials_and_numerals(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    found = {(f.tail, f.head.rstrip(",")) for f in result.line_breaks}
    assert ("Elizabeth", "II") in found      # a name split from its regnal numeral
    assert ("Davis", "Jr.") in found         # a name split from "Jr."
    assert ("W.", "H.") in found             # initials split from one another


def test_every_break_carries_a_verdict_for_every_rule(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    for brk in result.breaks:
        assert [f.rule for f in brk.findings][:9] == list(range(1, 10)), brk.display


def test_the_spreadsheet_has_the_tabs_and_page_numbers(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    path = report.write(result, tmp_path / "audit.xlsx")
    workbook = load_workbook(path)
    assert workbook.sheetnames == [
        "Summary", "Flagged Items", "All Instances", "Line Breaks (Rule 6)", "Rule Key",
    ]

    sheet = workbook["All Instances"]
    headers = [cell.value for cell in sheet[3]]
    assert headers[:6] == ["Page", "PDF page", "Break as set", "Word", "Type", "Verdict"]
    assert "R1" in headers and "R9" in headers

    pages = {row[0] for row in sheet.iter_rows(min_row=4, max_col=1, values_only=True)}
    assert pages == {"1", "2", "3", "4"}


def test_flagged_rows_all_state_a_reason(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    for brk in result.flagged + result.advisories:
        assert brk.reason, f"{brk.display} was flagged with no reason given"


def test_artifacts_are_not_counted_as_word_divisions(proof_pdf, tmp_path):
    result = run_audit(proof_pdf, tmp_path)
    counts = result.counts()
    assert counts["Extraction artifacts (no real break)"] == 1
    assert counts["Real word divisions checked"] == counts["End-of-line hyphens found"] - 1
    assert all(brk.worst is not Verdict.VIOLATION for brk in result.artifacts)
